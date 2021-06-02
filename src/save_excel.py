import os

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter
import numpy as np
from decimal import Decimal

dataset_list = ['cora', 'citeseer', 'cora_ml']
attack_list = ['meta', 'nettack', 'random']
defense_list = ['GCN', 'GCNSVD', 'Jaccard', 'ProGNN', 'RGCN', 'GAT', 'GNNGuard', 'CFS', 'CfsGAT', 'CfsRGCN']
defense_list1 = ['GCN', 'GCNSVD', 'Jaccard', 'ProGNN', 'RGCN', 'GAT', 'GNNGuard', 'Cfs', 'HGCN']

# 按行添加数据，参数：表格，开始的位置(i,j), 数据列表,按行或按列
def insert_data(ws, start_row, start_col, data, arg):
    if arg:
        # 按行
        col = start_col
        for i in data:
            ws.cell(start_row, col, i)
            col += 1
    else:
        row = start_row
        for i in data:
            ws.cell(row, start_col, i)
            row += 1

def create_sheet(t, name, defense_list):

    ts1 = t.create_sheet(name)

    ts1['A1'] = 'Dataset'
    ts1['B1'] = 'ptb rate%'

    insert_data(ts1, 1, 3, defense_list, 1)
    i = 2
    for dataset in dataset_list:
        ts1.merge_cells('A{}:A{}'.format(i, i + 5))
        ts1.cell(i, 1, dataset)
        i += 6

    if(name == 'meta'):
        for i in range(3):
            insert_data(ts1, i*6 + 2, 2, [j*5 for j in range(6)], 0)
    elif (name == 'nettack'):
        for i in range(3):
            insert_data(ts1, i*6 + 2, 2, [j for j in range(6)], 0)
    else:
        for i in range(3):
            insert_data(ts1, i*6 + 2, 2, [j*20 for j in range(6)], 0)
    return ts1

# 居中对齐设置列宽
def set_from_center(file,tableName):

    table = file.get_sheet_by_name(tableName)
    nrows = table.max_row  # 获得行数
    ncols = table.max_column
    for i in range(nrows):
        for j in range(ncols):
            table.cell(row=i+1, column=j+1).alignment = Alignment(horizontal='center', vertical='center')
            table.column_dimensions[get_column_letter(j + 1)].width = 15


def result(defense_list=defense_list, root='experiment/data/1/acc', root1='experiment/data/1/var',  save_dir='experiment/table/'):
    # defense_list = ['GCN', 'GCNSVD', 'Jaccard', 'ProGNN', 'RGCN', 'GAT', 'GNNGuard', 'HGCN', 'Cfs', 'Cs', 'CFSG']
    t = Workbook()

    # 写入数据
    for attack in attack_list:

        # 生成sheet
        ts = create_sheet(t, attack, defense_list)

        col_start = 3
        for defense in defense_list:
            row_start = 2
            for dataset in dataset_list:
                data = np.loadtxt('{}/{}_{}_{}.txt'.format(root, dataset, attack, defense))
                data_var = np.loadtxt('{}/{}_{}_{}.txt'.format(root1, dataset, attack, defense))
                data_com = [str(Decimal(data[i] * 100).quantize(Decimal("0.00"))) + "±" + str(
                    Decimal(data_var[i] * 100).quantize(Decimal("0.00"))) for i in range(6)]
                insert_data(ts, row_start, col_start, data_com, 0)
                row_start += 6
            col_start += 1
        set_from_center(t, attack)
    t.save('{}/result.xlsx'.format(save_dir))

def result_time(defense_list=defense_list1, root='experiment/data/4/',  save_dir='experiment/table/'):
    # defense_list = ['GCN', 'GCNSVD', 'Jaccard', 'ProGNN', 'RGCN', 'GAT', 'GNNGuard', 'Cfs', 'HGCN']
    t = Workbook()

    # 写入数据
    for attack in attack_list:

        # 生成sheet
        ts = create_sheet(t, attack, defense_list)

        col_start = 3
        for defense in defense_list:
            row_start = 2
            for dataset in dataset_list:
                data = np.loadtxt('{}/{}_{}_{}.txt'.format(root, dataset, attack, defense))
                data_com = [('%.2f' % data[i]) for i in range(6)]
                insert_data(ts, row_start, col_start, data_com, 0)
                row_start += 6
            col_start += 1
        set_from_center(t, attack)
    t.save('{}/result_time.xlsx'.format(save_dir))






